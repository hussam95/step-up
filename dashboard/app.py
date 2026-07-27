from __future__ import annotations

import io
import os
import re
from pathlib import Path
from functools import lru_cache

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from dash import Dash, Input, Output, State, dcc, html
from dash.exceptions import PreventUpdate


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "clean_data"
DOSAGE_FILE = DATA / "dosage_long.csv"

GREEN = "#2f7d6d"
GOLD = "#8a6f3d"
BLUE = "#4067a8"
RUST = "#b6573f"
INK = "#20242a"
TEMPLATE = "plotly_white"
SEASON_ORDER = {"Fall": 1, "Winter": 2, "Spring": 3}
STRICT_YEAR_SET = {"2022-23", "2023-24", "2024-25"}
STATUS_ORDER = ["Improved", "Regressed", "Stayed Same"]
STATUS_COLORS = {"Improved": GREEN, "Regressed": RUST, "Stayed Same": BLUE}
DOSAGE_BUCKET_ORDER = ["0%", "1-25%", "26-50%", "51-75%", "76-100%"]
DOSAGE_BUCKET_COLORS = {
    "0%": "#c8d1db",
    "1-25%": "#88a9d4",
    "26-50%": GOLD,
    "51-75%": GREEN,
    "76-100%": RUST,
}
OUTCOME_OPTIONS = [
    {"label": "STAR Reading", "value": "STAR Reading"},
    {"label": "STAR Math", "value": "STAR Math"},
    {"label": "CAASPP ELA", "value": "CAASPP ELA"},
    {"label": "CAASPP Math", "value": "CAASPP Math"},
    {"label": "Attendance", "value": "Attendance"},
]

def clean_name(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).lower().strip()
    if "," in text:
        last, first = text.split(",", 1)
        text = f"{first} {last}"
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def name_key(value) -> str:
    return clean_name(value).replace(" ", "")


def load_dosage_data() -> pd.DataFrame:
    if not DOSAGE_FILE.exists():
        return pd.DataFrame(
            columns=[
                "student_id",
                "student_name",
                "student_name_norm",
                "school_current",
                "school_group",
                "grade_current",
                "ethnicity_group",
                "student_group",
                "is_stepup",
                "program_count",
                "programs",
                "program_years",
                "dosage_sessions_attended",
                "dosage_sessions_possible",
                "dosage_rate",
                "dosage_bucket",
            ]
        )
    dosage = pd.read_csv(DOSAGE_FILE)
    for col in ["student_id", "student_name", "student_name_norm", "school_current", "school_group", "ethnicity_group", "student_group", "programs", "program_years", "dosage_bucket"]:
        if col in dosage.columns:
            dosage[col] = dosage[col].fillna("").astype(str)
    for col in ["grade_current", "program_count", "dosage_sessions_attended", "dosage_sessions_possible"]:
        if col in dosage.columns:
            dosage[col] = pd.to_numeric(dosage[col], errors="coerce")
    if "is_stepup" in dosage.columns:
        dosage["is_stepup"] = dosage["is_stepup"].astype(bool)
    if "dosage_rate" not in dosage.columns and {"dosage_sessions_attended", "dosage_sessions_possible"}.issubset(dosage.columns):
        dosage["dosage_rate"] = dosage["dosage_sessions_attended"] / dosage["dosage_sessions_possible"].replace({0: np.nan})
    dosage["dosage_rate"] = pd.to_numeric(dosage.get("dosage_rate"), errors="coerce").fillna(0.0)
    return dosage


def load_data() -> dict[str, pd.DataFrame]:
    students = pd.read_csv(DATA / "students.csv")
    star = pd.read_csv(DATA / "star_long.csv")
    attendance = pd.read_csv(DATA / "attendance_long.csv")
    caaspp = pd.read_csv(DATA / "caaspp_long.csv")
    growth = pd.read_csv(DATA / "star_growth_pairs.csv")
    availability = pd.read_csv(DATA / "availability_summary.csv")
    dosage = load_dosage_data()

    for df in [students, star, attendance, caaspp, growth]:
        df["student_id"] = df["student_id"].astype(str)
    if "student_id" in dosage.columns:
        dosage["student_id"] = dosage["student_id"].astype(str)

    students["student_label"] = (
        students["student_name"].fillna("Unknown").astype(str)
        + " | "
        + students["student_id"].astype(str)
        + " | "
        + students["school_current"].fillna("Unknown").astype(str)
    )
    return {
        "students": students,
        "star": star,
        "attendance": attendance,
        "caaspp": caaspp,
        "growth": growth,
        "availability": availability,
        "dosage": dosage,
    }


DS = load_data()


def opts(values):
    vals = sorted([v for v in pd.Series(values).dropna().unique()])
    return [{"label": str(v), "value": v} for v in vals]


def opts_from_frame(df: pd.DataFrame, label_col: str, value_col: str):
    if df.empty:
        return []
    return (
        df[[label_col, value_col]]
        .dropna()
        .drop_duplicates()
        .sort_values(label_col)
        .rename(columns={label_col: "label", value_col: "value"})
        .to_dict("records")
    )


def keep_valid_values(values, options):
    if values in (None, "", []):
        return []
    allowed = {str(option["value"]) for option in options}
    kept = [value for value in values if str(value) in allowed]
    return kept


def normalize_values(values):
    if values in (None, "", []):
        return ()
    if isinstance(values, (str, int, float, bool)):
        return (values,)
    return tuple(values)


def cache_key(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    return (
        mode,
        group,
        normalize_values(subjects),
        normalize_values(years),
        normalize_values(periods),
        normalize_values(schools),
        normalize_values(grades),
        normalize_values(ethnicities),
        normalize_values(student_ids),
    )


def empty_fig(title: str, message: str = "No data for this selection."):
    fig = go.Figure()
    fig.add_annotation(text=message, x=0.5, y=0.5, showarrow=False, font={"size": 16})
    fig.update_layout(template=TEMPLATE, title=title, xaxis={"visible": False}, yaxis={"visible": False})
    return fig


def polish(fig):
    fig.update_layout(
        template=TEMPLATE,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#ffffff",
        margin={"l": 46, "r": 24, "t": 76, "b": 48},
        title={"font": {"size": 19, "color": INK}, "x": 0.01, "xanchor": "left"},
        legend={
            "orientation": "h",
            "yanchor": "bottom",
            "y": 1.02,
            "xanchor": "right",
            "x": 1,
            "title": {"text": ""},
            "font": {"size": 12},
        },
        hoverlabel={"bgcolor": "#ffffff", "bordercolor": "#d8e0ea", "font": {"size": 12}},
        font={"family": "Inter, Segoe UI, Arial, sans-serif", "color": INK},
    )
    fig.update_xaxes(showgrid=False, zeroline=False, linecolor="#d8e0ea")
    fig.update_yaxes(gridcolor="#edf1f5", zeroline=False, linecolor="#d8e0ea")
    fig.for_each_annotation(lambda a: a.update(text=a.text.replace("subject=", "").replace("school_year=", "").replace("metric=", "")))
    return fig


def soften_axes(fig, x_title=None, y_title=None, hide_repeated_y=True):
    if x_title is not None:
        fig.update_xaxes(title_text=x_title)
    if y_title is not None:
        fig.update_yaxes(title_text=y_title)
    if hide_repeated_y:
        fig.for_each_yaxis(lambda axis: axis.update(title_text=""))
        if y_title:
            fig.update_yaxes(title_text=y_title, row=1, col=1)
    fig.update_yaxes(title_standoff=8, automargin=True)
    fig.update_xaxes(title_standoff=8, automargin=True)
    return fig


def kpi(label: str, value: str, note: str = ""):
    return html.Div(
        className="kpi-card",
        children=[html.Div(label, className="kpi-label"), html.Div(value, className="kpi-value"), html.Div(note, className="kpi-note")],
    )


def nav():
    links = [
        ("/", "Overview"),
        ("/summary", "Summary"),
        ("/star", "STAR Growth"),
        ("/attendance-caaspp", "Attendance and CAASPP"),
        ("/dosage", "Dosage vs Performance"),
        ("/coverage", "Data Coverage"),
    ]
    return html.Div(
        className="nav",
        children=[
            html.Div(className="brand", children=[html.Div("STEP UP", className="brand-title"), html.Div("Student Outcomes", className="brand-subtitle")]),
            html.Div(className="nav-links", children=[dcc.Link(label, href=href, className="nav-link") for href, label in links]),
        ],
    )


def filters():
    students = DS["students"]
    student_options = students.sort_values("student_label")[["student_label", "student_id"]].rename(
        columns={"student_label": "label", "student_id": "value"}
    )
    return html.Div(
        className="filters simple-filters",
        children=[
            html.Div(className="filter-control student-filter", children=[html.Label("Student / ID"), dcc.Dropdown(id="student-filter", options=student_options.to_dict("records"), value=[], multi=True, placeholder="Search name or District ID")]),
            html.Div(className="filter-control compact-radio", children=[html.Label("Student set"), dcc.RadioItems(id="analysis-mode", options=[{"label": "All available", "value": "all"}, {"label": "Complete comparison", "value": "strict"}], value="all", inline=True)]),
            html.Div(className="filter-control compact-radio", children=[html.Label("Group"), dcc.RadioItems(id="group-filter", options=[{"label": "All", "value": "all"}, {"label": "STEP UP", "value": "STEP UP"}, {"label": "Non-STEP UP", "value": "Non-STEP UP"}], value="all", inline=True)]),
            html.Div(className="filter-control", children=[html.Label("Subject"), dcc.Dropdown(id="subject-filter", options=[{"label": "Reading", "value": "Reading"}, {"label": "Math", "value": "Math"}], value=["Reading", "Math"], multi=True)]),
            html.Div(className="filter-control", children=[html.Label("Year"), dcc.Dropdown(id="year-filter", options=opts(DS["growth"]["school_year"]), value=[], multi=True, placeholder="All years")]),
            html.Div(className="filter-control", children=[html.Label("STAR period"), dcc.Dropdown(id="period-filter", options=opts(DS["growth"]["period"]), value=[], multi=True, placeholder="All periods")]),
            html.Div(className="filter-control", children=[html.Label("School"), dcc.Dropdown(id="school-filter", options=opts(students["school_group"]), value=[], multi=True, placeholder="All schools")]),
            html.Div(className="filter-control", children=[html.Label("Grade"), dcc.Dropdown(id="grade-filter", options=opts(students["grade_current"].dropna().astype(int)), value=[], multi=True, placeholder="All grades")]),
            html.Div(className="filter-control", children=[html.Label("Ethnicity"), dcc.Dropdown(id="ethnicity-filter", options=opts(students["ethnicity_group"]), value=[], multi=True, placeholder="All ethnicities")]),
            html.Div(className="filter-control", children=[html.Label("Dosage bucket"), dcc.Dropdown(id="dosage-bucket-filter", options=[{"label": bucket, "value": bucket} for bucket in DOSAGE_BUCKET_ORDER], value=[], multi=True, placeholder="All dosage buckets")]),
            html.Div(
                className="filter-note",
                style={"gridColumn": "1 / -1", "fontSize": "12px", "lineHeight": "1.4", "color": "#5f6b77", "paddingTop": "2px"},
                children="Strict mode stays on the 2022-25 comparison years for now. 2025-26 STAR and SIS attendance are available in All available mode; 2025-26 CAASPP is intentionally blank until it arrives.",
            ),
        ],
    )


def base_layout():
    return html.Div(
        [
            dcc.Location(id="url"),
            nav(),
            html.Div(
                className="app-shell",
                children=[
                    filters(),
                    dcc.Loading(
                        id="page-loading",
                        className="app-loading",
                        parent_className="app-loading-parent",
                        children=html.Div(id="page"),
                        type="dot",
                        color="#2f7d6d",
                        fullscreen=True,
                    ),
                ],
            ),
        ]
    )


def filter_frame(df: pd.DataFrame, group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids):
    out = df.copy()
    if mode == "strict" and "strict_comparison_ready" in out.columns:
        out = out[out["strict_comparison_ready"].astype(bool)]
    if mode == "strict":
        if "school_year_display" in out.columns:
            out = out[out["school_year_display"].isin(STRICT_YEAR_SET)]
        elif "school_year" in out.columns:
            out = out[out["school_year"].isin(STRICT_YEAR_SET)]
    if group != "all" and "student_group" in out.columns:
        out = out[out["student_group"].eq(group)]
    if subjects and "subject" in out.columns:
        subject_values = set(subjects)
        if "Reading" in subject_values:
            subject_values.add("ELA")
        out = out[out["subject"].isin(subject_values)]
    if years:
        if "school_year_display" in out.columns:
            out = out[out["school_year_display"].isin(years)]
        elif "school_year" in out.columns:
            out = out[out["school_year"].isin(years)]
    if periods and "period" in out.columns:
        out = out[out["period"].isin(periods)]
    if schools and "school_group" in out.columns:
        out = out[out["school_group"].isin(schools)]
    if grades and "grade_current" in out.columns:
        out = out[out["grade_current"].isin(grades)]
    if ethnicities and "ethnicity_group" in out.columns:
        out = out[out["ethnicity_group"].isin(ethnicities)]
    if student_ids and "student_id" in out.columns:
        out = out[out["student_id"].isin([str(x) for x in student_ids])]
    return out


@lru_cache(maxsize=128)
def cached_filtered(*key):
    mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids = key
    students = filter_frame(DS["students"], group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids)
    ids = set(students["student_id"])
    dosage = DS["dosage"][DS["dosage"]["student_id"].isin(ids)].copy()
    if years and len(dosage) and "program_years" in dosage.columns:
        year_values = {str(year) for year in years}
        dosage = dosage[dosage["program_years"].astype(str).apply(lambda text: any(year in text for year in year_values))]
    return {
        "students": students,
        "star": filter_frame(DS["star"][DS["star"]["student_id"].isin(ids)], group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids),
        "attendance": filter_frame(DS["attendance"][DS["attendance"]["student_id"].isin(ids)], group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids),
        "caaspp": filter_frame(DS["caaspp"][DS["caaspp"]["student_id"].isin(ids)], group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids),
        "growth": filter_frame(DS["growth"][DS["growth"]["student_id"].isin(ids)], group, mode, subjects, years, periods, schools, grades, ethnicities, student_ids),
        "dosage": dosage,
    }


def get_filtered(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    return cached_filtered(*cache_key(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids))


def filtered_availability(dfs):
    rows = []
    star = dfs["star"]
    attendance = dfs["attendance"]
    caaspp = dfs["caaspp"]
    if len(star):
        grouped = (
            star.groupby(["student_group", "school_year", "subject"])["student_id"]
            .nunique()
            .reset_index(name="students_available")
        )
        grouped["metric"] = "STAR"
        grouped = grouped.rename(columns={"school_year": "year"})
        rows.append(grouped[["student_group", "metric", "year", "subject", "students_available"]])
    if len(attendance):
        grouped = (
            attendance.groupby(["student_group", "school_year"])["student_id"]
            .nunique()
            .reset_index(name="students_available")
        )
        grouped["metric"] = "Attendance"
        grouped["subject"] = "Attendance"
        grouped = grouped.rename(columns={"school_year": "year"})
        rows.append(grouped[["student_group", "metric", "year", "subject", "students_available"]])
    if len(caaspp):
        grouped = (
            caaspp.groupby(["student_group", "school_year_display", "subject"])["student_id"]
            .nunique()
            .reset_index(name="students_available")
        )
        grouped["metric"] = "CAASPP"
        grouped = grouped.rename(columns={"school_year_display": "year"})
        rows.append(grouped[["student_group", "metric", "year", "subject", "students_available"]])
    if rows:
        return pd.concat(rows, ignore_index=True)
    return pd.DataFrame(columns=["student_group", "metric", "year", "subject", "students_available"])


def star_time_label(df: pd.DataFrame) -> pd.Series:
    return df["school_year"].astype(str) + " " + df["season"].astype(str)


def ordered_star_points(star: pd.DataFrame) -> pd.DataFrame:
    raw = star[star["value_type"].eq("score")].copy()
    if raw.empty:
        return raw
    raw["season_order"] = raw["season"].map(SEASON_ORDER).fillna(9)
    raw["time_order"] = raw["school_year"].str.slice(0, 4).astype(int) * 10 + raw["season_order"]
    raw["time_label"] = star_time_label(raw)
    return raw.sort_values(["student_id", "subject", "time_order"])


def first_last_pairs(df: pd.DataFrame, group_cols: list[str], metric: str, value_col: str = "value") -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    working = df.loc[df[value_col].notna(), group_cols + [value_col, "time_order", "time_label", "student_group", "student_name"]].sort_values(group_cols + ["time_order"])
    if working.empty:
        return pd.DataFrame()
    first = working.drop_duplicates(group_cols, keep="first").rename(
        columns={
            value_col: f"{value_col}_start",
            "time_order": "time_order_start",
            "time_label": "time_label_start",
            "student_group": "student_group_start",
            "student_name": "student_name_start",
        }
    )
    last = working.drop_duplicates(group_cols, keep="last").rename(
        columns={
            value_col: f"{value_col}_end",
            "time_order": "time_order_end",
            "time_label": "time_label_end",
            "student_group": "student_group_end",
            "student_name": "student_name_end",
        }
    )
    merged = first.merge(last, on=group_cols, suffixes=("_start", "_end"))
    merged = merged.loc[:, ~merged.columns.duplicated()].copy()
    merged = merged[merged["time_order_start"] != merged["time_order_end"]].copy()
    if merged.empty:
        return pd.DataFrame()
    merged["metric"] = metric
    merged["subject"] = merged["subject"] if "subject" in merged.columns else metric
    merged["student_group"] = merged["student_group_end"].combine_first(merged["student_group_start"])
    merged["student_name"] = merged["student_name_end"].combine_first(merged["student_name_start"])
    merged["start_label"] = merged["time_label_start"].fillna(merged["school_year_start"] if "school_year_start" in merged.columns else "")
    merged["end_label"] = merged["time_label_end"].fillna(merged["school_year_end"] if "school_year_end" in merged.columns else "")
    merged["start_value"] = merged[f"{value_col}_start"]
    merged["end_value"] = merged[f"{value_col}_end"]
    merged["change"] = merged["end_value"] - merged["start_value"]
    merged["change_status"] = "Stayed Same"
    merged.loc[merged["change"] > 0, "change_status"] = "Improved"
    merged.loc[merged["change"] < 0, "change_status"] = "Regressed"
    keep_cols = list(dict.fromkeys(group_cols + [
        "metric",
        "subject",
        "student_group",
        "student_name",
        "start_label",
        "end_label",
        "start_value",
        "end_value",
        "change",
        "change_status",
    ]))
    return merged[keep_cols]


def overview_outcome_pairs(dfs) -> pd.DataFrame:
    pieces = []
    star = ordered_star_points(dfs["star"])
    if len(star):
        pieces.append(first_last_pairs(star, ["student_id", "subject"], "STAR"))

    caaspp = dfs["caaspp"].copy()
    if len(caaspp):
        caaspp["time_order"] = caaspp["school_year_display"].astype(str).str.slice(0, 4).astype(int)
        caaspp["time_label"] = caaspp["school_year_display"]
        pieces.append(first_last_pairs(caaspp, ["student_id", "subject"], "CAASPP"))

    attendance = dfs["attendance"].copy()
    if len(attendance):
        attendance["measure_rank"] = attendance["measure"].map(
            {"SIS Reported Rate": 0, "Attendance rate": 1, "Full Day Rate": 2}
        ).fillna(9)
        attendance = attendance.sort_values("measure_rank").drop_duplicates(
            ["student_id", "school_year"], keep="first"
        )
        attendance["time_order"] = attendance["school_year"].astype(str).str.slice(0, 4).astype(int)
        attendance["time_label"] = attendance["school_year"]
        attendance["subject"] = "Attendance"
        pieces.append(first_last_pairs(attendance, ["student_id", "subject"], "Attendance"))

    if not pieces:
        return pd.DataFrame(
            columns=[
                "student_id",
                "metric",
                "subject",
                "student_group",
                "student_name",
                "start_label",
                "end_label",
                "start_value",
                "end_value",
                "change",
                "change_status",
            ]
        )
    out = pd.concat([p for p in pieces if len(p)], ignore_index=True)
    out["outcome"] = out.apply(
        lambda row: str(row["metric"]) if row["metric"] == "Attendance" else f"{row['metric']} {row['subject']}",
        axis=1,
    )
    out["outcome"] = out["outcome"].str.strip()
    return out


_BASE_PAIRS: pd.DataFrame | None = None


def base_pairs() -> pd.DataFrame:
    global _BASE_PAIRS
    if _BASE_PAIRS is None:
        _BASE_PAIRS = overview_outcome_pairs(DS)
    return _BASE_PAIRS


@lru_cache(maxsize=128)
def cached_overview_outcome_pairs(*key):
    dfs = get_filtered(*key)
    return overview_outcome_pairs(dfs)


def filter_pairs(pairs: pd.DataFrame, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    if pairs.empty:
        return pairs
    out = pairs
    if mode == "strict" and {"start_label", "end_label"}.issubset(out.columns):
        out = out[
            out["start_label"].astype(str).str[:7].isin(STRICT_YEAR_SET)
            & out["end_label"].astype(str).str[:7].isin(STRICT_YEAR_SET)
        ]
    if group != "all":
        out = out[out["student_group"].eq(group)]
    if subjects:
        subject_values = set(subjects)
        if "Reading" in subject_values:
            subject_values.add("ELA")
        out = out[out["subject"].isin(subject_values) | ((out["metric"].eq("Attendance")) & ("Attendance" in subject_values))]
    if years and "start_label" in out.columns:
        year_values = set(years)
        out = out[out["start_label"].astype(str).str[:7].isin(year_values) | out["end_label"].astype(str).str[:7].isin(year_values)]
    if student_ids:
        out = out[out["student_id"].isin([str(x) for x in student_ids])]
    return out


def outcome_kpi_value(pairs: pd.DataFrame, metric: str, subject: str | None = None, group: str | None = None) -> tuple[str, str]:
    sub = pairs[pairs["metric"].eq(metric)]
    if subject:
        sub = sub[sub["subject"].eq(subject)]
    if group:
        sub = sub[sub["student_group"].eq(group)]
    if sub.empty:
        return "0%", "No paired students"
    improved = int(sub["change_status"].eq("Improved").sum())
    total = len(sub)
    return f"{100 * improved / total:.1f}%", f"{improved:,} of {total:,} improved"


def primary_story_group(students: pd.DataFrame, requested_group: str | None = None) -> str | None:
    groups = set(students["student_group"].dropna())
    if requested_group in {"STEP UP", "Non-STEP UP"} and requested_group in groups:
        return requested_group
    if "STEP UP" in groups:
        return "STEP UP"
    if len(groups) == 1:
        return next(iter(groups))
    return None


def student_count_label(group: str | None, mode: str | None) -> tuple[str, str]:
    mode_note = "complete comparison students" if mode == "strict" else "students matching filters"
    if group == "STEP UP":
        return "STEP UP students in view", mode_note
    if group == "Non-STEP UP":
        return "Non-STEP UP students in view", mode_note
    return "Students in view", mode_note


def clean_snapshot_rows(pairs: pd.DataFrame, outcome: str | None) -> pd.DataFrame:
    if pairs.empty:
        return pd.DataFrame()
    selected = outcome or "STAR Reading"
    out = pairs[pairs["outcome"].eq(selected)].copy()
    out["change_display"] = out["change"].map(lambda x: f"{x:+.1f}")
    out["start_value_display"] = out["start_value"].map(lambda x: f"{x:,.1f}")
    out["end_value_display"] = out["end_value"].map(lambda x: f"{x:,.1f}")
    student_details = DS["students"][["student_id", "school_current", "grade_current", "ethnicity_group"]].drop_duplicates("student_id")
    out = out.merge(student_details, on="student_id", how="left")
    out["school_display"] = out["school_current"].fillna("Unknown")
    out["grade_display"] = out["grade_current"].map(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in {"", "nan"} else "Unknown"
    )
    out["ethnicity_display"] = out["ethnicity_group"].fillna("Unknown")
    return out.sort_values(["change_status", "change"], ascending=[True, False])


SNAPSHOT_COLUMN_HELP = {
    "Student": "The student in this row.",
    "School": "The current school for the student.",
    "Grade": "The current grade for the student.",
    "Ethnicity": "The student ethnicity group.",
    "Group": "The student group shown.",
    "Outcome": "The measure shown in this row.",
    "Start": "The first point being compared.",
    "End": "The later point being compared.",
    "Start Value": "The value at the start.",
    "End Value": "The value at the end.",
    "Change": "How much the value changed.",
    "Status": "Whether the value went up, went down, or stayed flat.",
}

COVERAGE_COLUMN_HELP = {
    "Metric": "The kind of data being counted.",
    "Year": "The school year shown here.",
    "Session": "The season or annual point.",
    "Subject": "The subject or measure.",
    "Group": "The student group shown.",
    "Students": "How many students have this data.",
    "Confidence": "How certain the year label is.",
}


def help_icon(text: str) -> html.Details:
    return html.Details(
        className="col-help",
        children=[
            html.Summary("i"),
            html.Div(text, className="col-help-popover"),
        ],
    )


def header_cell(label: str, help_text: str | None = None) -> html.Th:
    children = [html.Span(label)]
    if help_text:
        children.append(help_icon(help_text))
    return html.Th(html.Span(children, className="header-cell"))


def snapshot_table(df: pd.DataFrame) -> html.Div | html.Table:
    if df.empty:
        return html.Div(
            className="insight-band",
            children=[html.H3("No paired outcomes"), html.P("Select a broader group, year range, or outcome to show student-level changes.")],
        )
    working = df.copy()
    detail_cols = ["student_id", "school_current", "grade_current", "ethnicity_group"]
    if "school_display" not in working.columns or "grade_display" not in working.columns or "ethnicity_display" not in working.columns:
        if "student_id" in working.columns:
            student_details = DS["students"][detail_cols].drop_duplicates("student_id")
            working = working.merge(student_details, on="student_id", how="left")
        if "school_display" not in working.columns:
            school_source = working["school_current"] if "school_current" in working.columns else pd.Series([pd.NA] * len(working), index=working.index)
            working["school_display"] = school_source.fillna("Unknown")
        if "grade_display" not in working.columns:
            grade_source = working["grade_current"] if "grade_current" in working.columns else pd.Series([pd.NA] * len(working), index=working.index)
            working["grade_display"] = grade_source.map(
                lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in {"", "nan"} else "Unknown"
            )
        if "ethnicity_display" not in working.columns:
            ethnicity_source = working["ethnicity_group"] if "ethnicity_group" in working.columns else pd.Series([pd.NA] * len(working), index=working.index)
            working["ethnicity_display"] = ethnicity_source.fillna("Unknown")
    cols = [
        "student_name",
        "school_display",
        "grade_display",
        "ethnicity_display",
        "student_group",
        "outcome",
        "start_label",
        "end_label",
        "start_value_display",
        "end_value_display",
        "change_display",
        "change_status",
    ]
    labels = ["Student", "School", "Grade", "Ethnicity", "Group", "Outcome", "Start", "End", "Start Value", "End Value", "Change", "Status"]
    rows = []
    for _, row in working[cols].iterrows():
        status_class = f"status-{str(row['change_status']).lower().replace(' ', '-')}"
        rows.append(html.Tr([html.Td(str(row.get(c, ""))) for c in cols], className=status_class))
    return html.Table(
        className="data-table compact-table",
        children=[html.Thead(html.Tr([header_cell(label, SNAPSHOT_COLUMN_HELP.get(label)) for label in labels])), html.Tbody(rows)],
    )


def filtered_snapshot(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, outcome):
    pairs = filter_pairs(base_pairs(), mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids)
    return clean_snapshot_rows(pairs, outcome)


def page_cache_key(pathname, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, dosage_programs, dosage_buckets):
    return (pathname,) + cache_key(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids) + (
        normalize_values(dosage_programs),
        normalize_values(dosage_buckets),
    )


@lru_cache(maxsize=64)
def cached_render_from_key(key):
    pathname, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, dosage_programs, dosage_buckets = key
    dfs = get_filtered(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids)
    if pathname == "/star":
        return star_growth_page(dfs)
    if pathname == "/comparison":
        return comparison_page(dfs)
    if pathname == "/summary":
        return summary_page(dfs)
    if pathname == "/attendance-caaspp":
        return attendance_caaspp_page(dfs)
    if pathname == "/dosage":
        return dosage_page(dfs, dosage_programs, dosage_buckets)
    if pathname == "/students":
        return student_page(dfs)
    if pathname == "/coverage":
        return coverage_page(dfs)
    return overview(dfs, mode, group, years)


def snapshot_excel_bytes(df: pd.DataFrame) -> bytes:
    export = df[
        [
            "student_name",
            "school_display",
            "grade_display",
            "ethnicity_display",
            "student_group",
            "outcome",
            "start_label",
            "end_label",
            "start_value",
            "end_value",
            "change",
            "change_status",
        ]
    ].rename(
        columns={
            "student_name": "Student",
            "school_display": "School",
            "grade_display": "Grade",
            "ethnicity_display": "Ethnicity",
            "student_group": "Group",
            "outcome": "Outcome",
            "start_label": "Start",
            "end_label": "End",
            "start_value": "Start Value",
            "end_value": "End Value",
            "change": "Change",
            "change_status": "Status",
        }
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name="Outcome Snapshot", index=False)
        ws = writer.sheets["Outcome Snapshot"]
        from openpyxl.styles import Font, PatternFill

        fills = {
            "Improved": PatternFill("solid", fgColor="DCEFE8"),
            "Regressed": PatternFill("solid", fgColor="F8DEDA"),
            "Stayed Same": PatternFill("solid", fgColor="E9EDF3"),
        }
        header_fill = PatternFill("solid", fgColor="12313F")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        status_col = list(export.columns).index("Status") + 1
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row_idx, status_col).value
            fill = fills.get(status)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 34)
            ws.column_dimensions[column_cells[0].column_letter].width = width
        ws.freeze_panes = "A2"
    return output.getvalue()


SUMMARY_CATEGORY_ORDER = [
    "Outperformed non-STEP UP peers",
    "Improved, but below non-STEP UP peers",
    "Did not improve, below non-STEP UP peers",
]

SUMMARY_COLUMN_HELP = {
    "Metric": "The metric selected for this table.",
    "Student": "The student in this row.",
    "District ID": "The student District ID.",
    "School": "The current school for the student.",
    "Grade": "The current grade for the student.",
    "School Group": "The school group for the student.",
    "Ethnicity": "The student ethnicity group.",
    "Intensity": "The STEP UP intensity level.",
    "Year": "The school year being shown.",
    "Period": "The STAR period or annual measure.",
    "Start": "The first point being compared.",
    "End": "The later point being compared.",
    "Start Value": "The value at the start.",
    "End Value": "The value at the end.",
    "Change": "How much the value changed.",
    "Benchmark": "The average change for non-STEP UP students in the same year and metric.",
    "Delta vs Benchmark": "How far this student is above or below the non-STEP UP average.",
    "Summary Category": "Whether the student outperformed non-STEP UP peers, improved below them, or did not improve below them.",
    "Status": "Whether the value went up, went down, or stayed flat.",
}


def summary_metric_label(metric: str) -> str:
    return str(metric or "STAR Reading")


def summary_filtered_dfs(mode, group, years, schools, grades, ethnicities, student_ids):
    return {
        "students": filter_frame(DS["students"], group, mode, None, years, None, schools, grades, ethnicities, student_ids),
        "star": filter_frame(DS["star"], group, mode, None, years, None, schools, grades, ethnicities, student_ids),
        "attendance": filter_frame(DS["attendance"], group, mode, None, years, None, schools, grades, ethnicities, student_ids),
        "caaspp": filter_frame(DS["caaspp"], group, mode, None, years, None, schools, grades, ethnicities, student_ids),
    }


def summary_merge_student_details(df: pd.DataFrame, students: pd.DataFrame) -> pd.DataFrame:
    detail_cols = ["student_id", "school_current", "school_group", "grade_current", "ethnicity_group", "intervention_intensity"]
    return df.merge(students[detail_cols], on="student_id", how="left")


def summary_metric_rows(dfs: dict[str, pd.DataFrame], metric: str) -> pd.DataFrame:
    metric = summary_metric_label(metric)
    students = dfs["students"]
    pairs = overview_outcome_pairs(dfs)
    if pairs.empty:
        return pairs
    if metric.startswith("STAR "):
        pairs = pairs[pairs["metric"].eq("STAR")].copy()
        subject = metric.split(" ", 1)[1]
        pairs = pairs[pairs["subject"].eq(subject)].copy()
        bench_group = ["year", "subject"]
        pairs["year"] = pairs["start_label"].astype(str).str.extract(r"^([^ ]+)")[0]
    else:
        pairs = pairs[pairs["metric"].eq("CAASPP" if metric.startswith("CAASPP ") else "Attendance")].copy()
        if metric.startswith("CAASPP "):
            subject = metric.split(" ", 1)[1]
            pairs = pairs[pairs["subject"].eq(subject)].copy()
        bench_group = ["year", "subject"] if metric.startswith("CAASPP ") else ["year"]
        pairs["year"] = pairs.apply(
            lambda row: row["start_label"] if str(row["start_label"]) == str(row["end_label"]) else f"{row['start_label']} → {row['end_label']}",
            axis=1,
        )
    if pairs.empty:
        return pairs
    pairs = summary_merge_student_details(pairs, students)
    pairs["period_label"] = "Annual"
    bench = (
        pairs[pairs["student_group"].eq("Non-STEP UP")]
        .groupby(bench_group)["change"]
        .mean()
        .reset_index(name="benchmark")
    )
    rows = pairs.merge(bench, on=bench_group, how="left")
    rows["summary_category"] = "Did not improve, below non-STEP UP peers"
    rows.loc[rows["change"] > rows["benchmark"], "summary_category"] = "Outperformed non-STEP UP peers"
    rows.loc[
        (rows["change"] > 0) & (rows["change"] <= rows["benchmark"]),
        "summary_category",
    ] = "Improved, but below non-STEP UP peers"
    rows["metric"] = metric

    rows = rows[rows["student_group"].eq("STEP UP")].copy()
    if rows.empty:
        return rows
    rows["delta_vs_benchmark"] = rows["change"] - rows["benchmark"]
    rows["start_value_display"] = rows["start_value"].map(lambda x: f"{x:,.1f}")
    rows["end_value_display"] = rows["end_value"].map(lambda x: f"{x:,.1f}")
    rows["change_display"] = rows["change"].map(lambda x: f"{x:+.1f}")
    rows["benchmark_display"] = rows["benchmark"].map(lambda x: f"{x:+.1f}")
    rows["delta_display"] = rows["delta_vs_benchmark"].map(lambda x: f"{x:+.1f}")
    school_current = rows["school_current"] if "school_current" in rows.columns else pd.Series([pd.NA] * len(rows), index=rows.index)
    school_source = rows["school"] if "school" in rows.columns else pd.Series([pd.NA] * len(rows), index=rows.index)
    grade_current = rows["grade_current"] if "grade_current" in rows.columns else pd.Series([pd.NA] * len(rows), index=rows.index)
    grade_source = rows["grade_num"] if "grade_num" in rows.columns else pd.Series([pd.NA] * len(rows), index=rows.index)
    rows["school_display"] = school_current.fillna(school_source).fillna("Unknown")
    rows["grade_display"] = grade_current.fillna(grade_source).map(
        lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in {"", "nan"} else "Unknown"
    )
    rows["school_group"] = rows["school_group"].fillna("Unknown") if "school_group" in rows.columns else "Unknown"
    rows["ethnicity_group"] = rows["ethnicity_group"].fillna("Unknown") if "ethnicity_group" in rows.columns else "Unknown"
    rows["intervention_intensity"] = rows["intervention_intensity"].fillna("Unknown") if "intervention_intensity" in rows.columns else "Unknown"
    rows["start_display"] = rows["start_label"].fillna("")
    rows["end_display"] = rows["end_label"].fillna("")
    rows = rows.sort_values(["summary_category", "change"], ascending=[True, False])
    return rows


def summary_table(df: pd.DataFrame) -> html.Div | html.Table:
    if df.empty:
        return html.Div(
            className="insight-band",
            children=[html.H3("No STEP UP rows"), html.P("Try a different year, group, or metric to show the summary table.")],
        )
    cols = [
        "metric",
        "student_name",
        "student_id",
        "school_display",
        "grade_display",
        "school_group",
        "ethnicity_group",
        "intervention_intensity",
        "year",
        "period_label",
        "start_display",
        "end_display",
        "start_value_display",
        "end_value_display",
        "change_display",
        "benchmark_display",
        "delta_display",
        "summary_category",
        "change_status",
    ]
    labels = [
        "Metric",
        "Student",
        "District ID",
        "School",
        "Grade",
        "School Group",
        "Ethnicity",
        "Intensity",
        "Year",
        "Period",
        "Start",
        "End",
        "Start Value",
        "End Value",
        "Change",
        "Benchmark",
        "Delta vs Benchmark",
        "Summary Category",
        "Status",
    ]
    rows = []
    for _, row in df[cols].iterrows():
        status_class = f"status-{str(row['change_status']).lower().replace(' ', '-')}"
        rows.append(html.Tr([html.Td(str(row.get(c, ""))) for c in cols], className=status_class))
    return html.Table(
        className="data-table compact-table",
        children=[html.Thead(html.Tr([header_cell(label, SUMMARY_COLUMN_HELP.get(label)) for label in labels])), html.Tbody(rows)],
    )


def summary_excel_bytes(df: pd.DataFrame) -> bytes:
    export = df[
        [
            "metric",
            "student_name",
            "student_id",
            "school_display",
            "grade_display",
            "school_group",
            "ethnicity_group",
            "intervention_intensity",
            "year",
            "period_label",
            "start_display",
            "end_display",
            "start_value_display",
            "end_value_display",
            "change_display",
            "benchmark_display",
            "delta_display",
            "summary_category",
            "change_status",
        ]
    ].rename(
        columns={
            "metric": "Metric",
            "student_name": "Student",
            "student_id": "District ID",
            "school_display": "School",
            "grade_display": "Grade",
            "school_group": "School Group",
            "ethnicity_group": "Ethnicity",
            "intervention_intensity": "Intensity",
            "year": "Year",
            "period_label": "Period",
            "start_display": "Start",
            "end_display": "End",
            "start_value_display": "Start Value",
            "end_value_display": "End Value",
            "change_display": "Change",
            "benchmark_display": "Benchmark",
            "delta_display": "Delta vs Benchmark",
            "summary_category": "Summary Category",
            "change_status": "Status",
        }
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        from openpyxl.styles import Font, PatternFill

        fills = {
            "Improved": PatternFill("solid", fgColor="DCEFE8"),
            "Regressed": PatternFill("solid", fgColor="F8DEDA"),
            "Stayed Same": PatternFill("solid", fgColor="E9EDF3"),
        }
        header_fill = PatternFill("solid", fgColor="12313F")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        status_col = list(export.columns).index("Status") + 1
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row_idx, status_col).value
            fill = fills.get(status)
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 36)
            ws.column_dimensions[column_cells[0].column_letter].width = width
        ws.freeze_panes = "A2"
    return output.getvalue()


def summary_cards(df: pd.DataFrame, metric: str, years) -> html.Div:
    metric_label = summary_metric_label(metric)
    selected_years = ", ".join(map(str, years)) if years else "all available years"
    if df.empty:
        return html.Div(
            className="insight-band",
            children=[
                html.H3("Summary snapshot"),
                html.P(f"No STEP UP rows are available for {metric_label} in {selected_years}. Try widening the filters."),
            ],
        )
    outperformed = int(df["summary_category"].eq("Outperformed non-STEP UP peers").sum())
    improved = int(df["summary_category"].eq("Improved, but below non-STEP UP peers").sum())
    not_improved = int(df["summary_category"].eq("Did not improve, below non-STEP UP peers").sum())
    return html.Div(
        children=[
            html.Div(
                className="kpi-grid",
                children=[
                    kpi("STEP UP students", f"{len(df):,}", f"{metric_label} for {selected_years}"),
                    kpi("Outperformed peers", f"{outperformed:,}", "Beat the non-STEP UP average"),
                    kpi("Improved below peers", f"{improved:,}", "Went up, but not above the average"),
                    kpi("Not improved below peers", f"{not_improved:,}", "Did not improve versus the average"),
                ],
            ),
            html.Div(
                className="insight-band",
                children=[
                    html.H3("How to read this table"),
                    html.P(
                        "Rows are STEP UP students only. The Metric column shows which measure you selected. "
                        "Outperformed means the student did better than the non-STEP UP average. "
                        "Improved, but below means the student got better, but not as much as the non-STEP UP average. "
                        "Did not improve means the student did not show a gain."
                    ),
                ],
            ),
            html.Div(summary_table(df), className="table-scroll"),
        ]
    )


def outcome_direction_summary(pairs: pd.DataFrame) -> pd.DataFrame:
    if pairs.empty:
        return pd.DataFrame()
    summary = (
        pairs.groupby(["outcome", "student_group", "change_status"])
        .size()
        .reset_index(name="students")
    )
    totals = summary.groupby(["outcome", "student_group"])["students"].transform("sum")
    summary["percent"] = 100 * summary["students"] / totals
    return summary


def star_trend_figure(star: pd.DataFrame, title: str):
    raw = ordered_star_points(star)
    if raw.empty:
        return empty_fig(title)
    trend = (
        raw.groupby(["time_order", "time_label", "subject", "student_group"])["value"]
        .mean()
        .reset_index()
        .sort_values("time_order")
    )
    fig = px.line(
        trend,
        x="time_label",
        y="value",
        color="student_group",
        facet_col="subject",
        markers=True,
        title=title,
        template=TEMPLATE,
        color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
        labels={"time_label": "Time", "value": "Avg STAR", "student_group": "", "subject": "Subject"},
    )
    fig = polish(fig)
    return soften_axes(fig, x_title="", y_title="Avg STAR")


def availability_figures(dfs):
    avail = filtered_availability(dfs)
    if len(avail):
        avail_plot = avail.copy()
        avail_plot["measure_subject"] = avail_plot.apply(
            lambda r: r["metric"] if r["metric"] == "Attendance" else f"{r['metric']} {r['subject']}",
            axis=1,
        )
        fig_avail = px.bar(
            avail_plot,
            x="year",
            y="students_available",
            color="student_group",
            facet_col="measure_subject",
            facet_col_wrap=3,
            barmode="group",
            title="Students with available data by group, measure, and subject",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"year": "Year", "students_available": "Students with data", "student_group": "", "measure_subject": ""},
        )
        fig_avail.update_yaxes(matches=None)
        fig_avail = soften_axes(polish(fig_avail), x_title="", y_title="Students")
    else:
        fig_avail = empty_fig("Students with available data")

    star = dfs["star"]
    if len(star):
        star_summary = (
            star.groupby(["school_year", "season", "subject", "student_group"])["student_id"]
            .nunique()
            .reset_index(name="students_available")
        )
        fig_star_avail = px.bar(
            star_summary,
            x="season",
            y="students_available",
            color="student_group",
            facet_row="school_year",
            facet_col="subject",
            barmode="group",
            category_orders={"season": ["Fall", "Winter", "Spring"]},
            title="STAR availability by season: Reading and Math",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"season": "Season", "students_available": "Students with STAR data", "student_group": "", "subject": "Subject", "school_year": "Year"},
        )
        fig_star_avail.update_yaxes(matches=None)
        fig_star_avail = soften_axes(polish(fig_star_avail), x_title="", y_title="Students")
    else:
        fig_star_avail = empty_fig("STAR availability by season")
    return fig_star_avail, fig_avail


def growth_kpis(growth: pd.DataFrame):
    n = growth["student_id"].nunique()
    if len(growth) == 0:
        return html.Div(className="kpi-grid", children=[kpi("Students with start/end values", "0", "Selected STAR period")])
    counts = growth["change_status"].value_counts()
    improved = int(counts.get("Improved", 0))
    regressed = int(counts.get("Regressed", 0))
    same = int(counts.get("Stayed Same", 0))
    total_pairs = improved + regressed + same
    pct = lambda x: "0%" if total_pairs == 0 else f"{100*x/total_pairs:.1f}%"
    return html.Div(
        className="kpi-grid",
        children=[
            kpi("Students with start/end values", f"{n:,}", "Unique students in selected STAR comparison"),
            kpi("Improved", pct(improved), f"{improved:,} student-period comparisons"),
            kpi("Regressed", pct(regressed), f"{regressed:,} student-period comparisons"),
            kpi("Stayed same", pct(same), f"{same:,} student-period comparisons"),
        ],
    )


def page(title: str, subtitle: str, children):
    return html.Div(className="page", children=[html.Div(className="page-header", children=[html.H1(title), html.P(subtitle)]), children])


def loading_indicator():
    return html.Div(
        className="loading-overlay",
        children=html.Div(
            className="loading-card",
            children=[
                html.Div(className="loading-mark", children=[html.Span(), html.Span(), html.Span()]),
                html.Div(
                    className="loading-copy",
                    children=[
                        html.Div("Loading dashboard", className="loading-title"),
                        html.Div("Refreshing tabs, filters, and records", className="loading-subtitle"),
                    ],
                ),
            ],
        ),
    )


def error_panel(message: str):
    return html.Div(
        className="insight-band",
        children=[
            html.H3("Dashboard load error"),
            html.P(message),
        ],
    )


def overview(dfs, mode="all", requested_group="all", selected_years=None):
    students = dfs["students"]
    pairs = filter_pairs(base_pairs(), mode, requested_group, None, None, None, None, None, None, None)
    story_group = primary_story_group(students, requested_group)
    story_prefix = story_group if story_group else "Selected"
    count_group = requested_group if requested_group in {"STEP UP", "Non-STEP UP"} else None
    count_label, count_note = student_count_label(count_group, mode)
    if selected_years:
        metric_ids = set()
        for key in ["star", "attendance", "caaspp", "growth"]:
            if key in dfs and len(dfs[key]):
                metric_ids.update(dfs[key]["student_id"].astype(str))
        visible_students = set(students["student_id"].astype(str)) & metric_ids
        student_count = len(visible_students)
        count_note = f"students with {', '.join(map(str, selected_years))} data"
    else:
        student_count = students["student_id"].nunique()
    star_reading = outcome_kpi_value(pairs, "STAR", "Reading", story_group)
    star_math = outcome_kpi_value(pairs, "STAR", "Math", story_group)
    caaspp_ela = outcome_kpi_value(pairs, "CAASPP", "ELA", story_group)
    caaspp_math = outcome_kpi_value(pairs, "CAASPP", "Math", story_group)
    attendance = outcome_kpi_value(pairs, "Attendance", "Attendance", story_group)
    cards = html.Div(
        className="kpi-grid",
        children=[
            kpi(count_label, f"{student_count:,}", count_note),
            kpi(f"{story_prefix} STAR Reading improved", star_reading[0], star_reading[1]),
            kpi(f"{story_prefix} STAR Math improved", star_math[0], star_math[1]),
            kpi(f"{story_prefix} Attendance improved", attendance[0], attendance[1]),
            kpi(f"{story_prefix} CAASPP ELA improved", caaspp_ela[0], caaspp_ela[1]),
            kpi(f"{story_prefix} CAASPP Math improved", caaspp_math[0], caaspp_math[1]),
        ],
    )

    direction = outcome_direction_summary(pairs)
    if len(direction):
        fig_direction = px.bar(
            direction,
            x="outcome",
            y="percent",
            color="change_status",
            facet_col="student_group",
            barmode="stack",
            category_orders={"change_status": STATUS_ORDER},
            title="Outcome direction from earliest to latest available point",
            template=TEMPLATE,
            color_discrete_map=STATUS_COLORS,
            labels={"outcome": "Outcome", "percent": "% of paired students", "change_status": "", "student_group": ""},
            hover_data={"students": True, "percent": ":.1f"},
        )
        fig_direction.update_yaxes(range=[0, 100])
        fig_direction = soften_axes(polish(fig_direction), x_title="", y_title="% students")
    else:
        fig_direction = empty_fig("Outcome direction from earliest to latest available point")

    fig_star_story = star_trend_figure(dfs["star"], "STAR score journey over available seasons")

    caaspp = dfs["caaspp"]
    if len(caaspp):
        ca_trend = caaspp.groupby(["school_year_display", "subject", "student_group"])["value"].mean().reset_index()
        fig_ca = px.line(
            ca_trend,
            x="school_year_display",
            y="value",
            color="student_group",
            facet_col="subject",
            markers=True,
            title="CAASPP movement where two years are available",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"school_year_display": "Year", "value": "Avg CAASPP", "student_group": "", "subject": "Subject"},
        )
        fig_ca = soften_axes(polish(fig_ca), x_title="", y_title="Avg CAASPP")
    else:
        fig_ca = empty_fig("CAASPP movement where two years are available")

    att = dfs["attendance"]
    if len(att):
        att = att.copy()
        att["measure_rank"] = att["measure"].map({"SIS Reported Rate": 0, "Attendance rate": 1, "Full Day Rate": 2}).fillna(9)
        att = att.sort_values("measure_rank").drop_duplicates(["student_id", "school_year"], keep="first")
        att_year = (
            att.groupby(["school_year", "student_group"])["value"]
            .mean()
            .reset_index()
        )
        fig_att = px.line(
            att_year,
            x="school_year",
            y="value",
            color="student_group",
            markers=True,
            title="Attendance movement across available years",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"school_year": "Year", "value": "Avg attendance", "student_group": ""},
        )
        fig_att.update_yaxes(range=[0, 100])
        fig_att = soften_axes(polish(fig_att), x_title="", y_title="Avg attendance", hide_repeated_y=False)
    else:
        fig_att = empty_fig("Attendance movement across available years")

    note = html.Div(
        className="insight-band",
        children=[
            html.H3("How to read this page"),
            html.P("Overview comparisons use each student's earliest and latest selected point for STAR, CAASPP, and attendance. Seasonal STAR detail and raw data availability live on the STAR Growth and Data Coverage tabs."),
        ],
    )
    return page(
        "STEP UP Student Story",
        "Earliest-to-latest movement across STAR, CAASPP, and attendance, with STEP UP students kept in view even when some metrics are missing.",
        html.Div(
            [
                cards,
                note,
                dcc.Graph(figure=fig_direction, className="chart wide"),
                dcc.Graph(figure=fig_star_story, className="chart wide"),
                html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca), dcc.Graph(figure=fig_att)]),
                html.Div(
                    className="section-toolbar",
                    children=[
                        html.H2("Student-level outcome snapshot"),
                        html.Div(
                            className="snapshot-actions",
                            children=[
                                dcc.Dropdown(id="overview-outcome-filter", options=OUTCOME_OPTIONS, value="STAR Reading", clearable=False),
                                html.Button("Download Excel", id="download-snapshot-button", className="download-button", n_clicks=0),
                                dcc.Download(id="download-snapshot"),
                            ],
                        ),
                    ],
                ),
                html.Div(id="overview-snapshot-table", className="table-scroll"),
            ]
        ),
    )


def summary_page(dfs):
    controls = html.Div(
        className="section-toolbar",
        children=[
            html.H2("Summary metric"),
            html.Div(
                className="snapshot-actions",
                children=[
                    html.Div(
                        className="filter-control",
                        children=[
                            html.Label("Metric"),
                            dcc.Dropdown(
                                id="summary-metric-filter",
                                options=OUTCOME_OPTIONS,
                                value="STAR Reading",
                                clearable=False,
                                persistence=True,
                                persistence_type="session",
                            ),
                        ],
                    ),
                    html.Button("Download Excel", id="download-summary-button", className="download-button", n_clicks=0),
                    dcc.Download(id="download-summary"),
                ],
            ),
        ],
    )
    note = html.Div(
        className="insight-band",
        children=[
            html.H3("How to read this page"),
            html.P(
                "Use the Year filter to focus on one year, then choose a metric. "
                "The table shows STEP UP students only. "
                "Outperformed means the student did better than the non-STEP UP average. "
                "Improved, but below means the student improved, but not enough to beat the non-STEP UP average. "
                "Did not improve means the student did not show a gain."
            ),
        ],
    )
    return page(
        "Summary",
        "STEP UP students compared with non-STEP UP peers by metric and year.",
        html.Div([controls, note, html.Div(id="summary-content")]),
    )


def star_growth_page(dfs):
    star, growth = dfs["star"], dfs["growth"]
    star_pairs = overview_outcome_pairs({**dfs, "attendance": dfs["attendance"].iloc[0:0], "caaspp": dfs["caaspp"].iloc[0:0]})
    star_pairs = star_pairs[star_pairs["metric"].eq("STAR")]
    story_group = primary_story_group(dfs["students"])
    reading = outcome_kpi_value(star_pairs, "STAR", "Reading", story_group)
    math = outcome_kpi_value(star_pairs, "STAR", "Math", story_group)
    paired_students = star_pairs["student_id"].nunique() if len(star_pairs) else 0

    if len(growth):
        period_order = ["Fall to Winter", "Winter to Spring", "Fall to Spring"]
        period_status = (
            growth.groupby(["student_group", "subject", "period", "change_status"])
            .size()
            .reset_index(name="students")
        )
        period_status["total"] = period_status.groupby(["student_group", "subject", "period"])["students"].transform("sum")
        period_status["percent"] = 100 * period_status["students"] / period_status["total"]
        improved_periods = period_status[period_status["change_status"].eq("Improved")].copy()
        if len(improved_periods):
            best = improved_periods.sort_values("percent", ascending=False).iloc[0]
            best_note = f"{best['student_group']} {best['subject']} {best['period']}"
            best_value = f"{best['percent']:.1f}%"
        else:
            best_value, best_note = "0%", "No improved period"

        fig_period = px.bar(
            period_status,
            x="period",
            y="percent",
            color="change_status",
            facet_row="student_group",
            facet_col="subject",
            barmode="stack",
            category_orders={"change_status": STATUS_ORDER},
            title="Where STAR students improved, regressed, or stayed flat across all available years",
            template=TEMPLATE,
            color_discrete_map=STATUS_COLORS,
            labels={"period": "STAR period", "percent": "% of paired students", "change_status": "", "student_group": "", "subject": ""},
            hover_data={"students": True, "percent": ":.1f"},
        )
        fig_period.update_yaxes(range=[0, 100], matches=None)
        fig_period = soften_axes(polish(fig_period), x_title="", y_title="% students")

        fig_dist = px.box(
            growth,
            x="period",
            y="change",
            color="student_group",
            facet_col="subject",
            points=False,
            title="Distribution of STAR score changes by period across all available years",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            category_orders={"period": period_order},
            labels={"period": "STAR period", "change": "Score change", "student_group": "", "subject": ""},
        )
        fig_dist = soften_axes(polish(fig_dist), x_title="", y_title="Score change")

        avg_change = growth.groupby(["student_group", "school_year", "subject", "period"])["change"].mean().reset_index()
        fig_avg = px.bar(
            avg_change,
            x="period",
            y="change",
            color="student_group",
            facet_row="school_year",
            facet_col="subject",
            barmode="group",
            title="Average STAR change by year and period",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"period": "STAR period", "change": "Avg change", "student_group": "", "school_year": "Year", "subject": ""},
        )
        fig_avg.update_yaxes(matches=None)
        fig_avg = soften_axes(polish(fig_avg), x_title="", y_title="Avg change")
    else:
        best_value, best_note = "0%", "No STAR period pairs"
        fig_period = empty_fig("Where STAR students improved, regressed, or stayed flat across all available years")
        fig_dist = empty_fig("Distribution of STAR score changes by period across all available years")
        fig_avg = empty_fig("Average STAR change by year and period")

    if len(star_pairs):
        table_df = star_pairs.copy()
        table_df["change_display"] = table_df["change"].map(lambda x: f"{x:+.1f}")
        table_df["start_value_display"] = table_df["start_value"].map(lambda x: f"{x:,.1f}")
        table_df["end_value_display"] = table_df["end_value"].map(lambda x: f"{x:,.1f}")
        table_df = table_df.sort_values(["change_status", "change"], ascending=[True, False])
        table = snapshot_table(table_df)
    else:
        table = snapshot_table(pd.DataFrame())

    cards = html.Div(
        className="kpi-grid",
        children=[
            kpi("Students with STAR pairs", f"{paired_students:,}", "Earliest-to-latest STAR records"),
            kpi("STAR Reading improved", reading[0], reading[1]),
            kpi("STAR Math improved", math[0], math[1]),
            kpi("Strongest improvement period", best_value, best_note),
        ],
    )
    note = html.Div(
        className="insight-band",
        children=[
            html.H3("What this page adds"),
            html.P("The first chart combines all available years and shows the share of students who improved, regressed, or stayed flat by STAR period. The middle chart shows the spread of score changes by STAR period across all available years. The year-by-year chart below separates changes by school year so you can compare trends over time."),
        ],
    )
    return page(
        "STAR Growth",
        "Period-by-period STAR growth details behind the overview story.",
        html.Div([cards, note, dcc.Graph(figure=fig_period, className="chart wide"), html.Div(className="two-col", children=[dcc.Graph(figure=fig_dist), dcc.Graph(figure=fig_avg)]), html.H2("STAR student change details"), html.Div(table, className="table-scroll")]),
    )


def comparison_page(dfs):
    star, caaspp, attendance = dfs["star"], dfs["caaspp"], dfs["attendance"]
    star_use = star[star["value_type"].isin(["score", "benchmark_numeric"])]
    if len(star_use):
        star_summary = star_use.groupby(["student_group", "school_year", "subject"])["value"].mean().reset_index()
        fig_star = px.bar(
            star_summary,
            x="school_year",
            y="value",
            color="student_group",
            facet_col="subject",
            barmode="group",
            title="Average STAR by group",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"school_year": "Year", "value": "Average STAR value", "student_group": "", "subject": "Subject"},
        )
        fig_star = polish(fig_star)
        fig_star = soften_axes(fig_star, x_title="", y_title="Avg STAR")
    else:
        fig_star = empty_fig("Average STAR by group")
    if len(caaspp):
        ca_summary = caaspp.groupby(["student_group", "school_year_display", "subject"])["value"].mean().reset_index()
        fig_ca = px.bar(
            ca_summary,
            x="school_year_display",
            y="value",
            color="student_group",
            facet_col="subject",
            barmode="group",
            title="Average CAASPP by group",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"school_year_display": "Year", "value": "Average CAASPP score", "student_group": "", "subject": "Subject"},
        )
        fig_ca = polish(fig_ca)
        fig_ca = soften_axes(fig_ca, x_title="", y_title="Avg CAASPP")
    else:
        fig_ca = empty_fig("Average CAASPP by group")
    if len(attendance):
        att_summary = attendance.groupby(["student_group", "school_year"])["value"].mean().reset_index()
        fig_att = px.bar(
            att_summary,
            x="school_year",
            y="value",
            color="student_group",
            barmode="group",
            title="Average attendance by group",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"school_year": "Year", "value": "Average attendance rate", "student_group": ""},
        )
        fig_att.update_yaxes(range=[0, 100])
        fig_att = polish(fig_att)
        fig_att = soften_axes(fig_att, x_title="", y_title="Avg attendance", hide_repeated_y=False)
    else:
        fig_att = empty_fig("Average attendance by group")
    return page("STEP UP vs Non-STEP UP", "Use all available data or switch to strict comparable data in the Mode filter.", html.Div([dcc.Graph(figure=fig_star, className="chart wide"), html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca), dcc.Graph(figure=fig_att)])]))


def attendance_caaspp_page(dfs):
    attendance, caaspp = dfs["attendance"], dfs["caaspp"]
    pairs = overview_outcome_pairs(dfs)
    pairs = pairs[pairs["metric"].isin(["CAASPP", "Attendance"])]
    story_group = primary_story_group(dfs["students"])
    ca_ela = outcome_kpi_value(pairs, "CAASPP", "ELA", story_group)
    ca_math = outcome_kpi_value(pairs, "CAASPP", "Math", story_group)
    att_kpi = outcome_kpi_value(pairs, "Attendance", "Attendance", story_group)
    paired_students = pairs["student_id"].nunique() if len(pairs) else 0

    direction = outcome_direction_summary(pairs)
    if len(direction):
        fig_direction = px.bar(
            direction,
            x="outcome",
            y="percent",
            color="change_status",
            facet_col="student_group",
            barmode="stack",
            category_orders={"change_status": STATUS_ORDER},
            title="Annual outcome direction for CAASPP and attendance",
            template=TEMPLATE,
            color_discrete_map=STATUS_COLORS,
            labels={"outcome": "Outcome", "percent": "% of paired students", "change_status": "", "student_group": ""},
            hover_data={"students": True, "percent": ":.1f"},
        )
        fig_direction.update_yaxes(range=[0, 100])
        fig_direction = soften_axes(polish(fig_direction), x_title="", y_title="% students")
    else:
        fig_direction = empty_fig("Annual outcome direction for CAASPP and attendance")

    ca_pairs = pairs[pairs["metric"].eq("CAASPP")]
    if len(ca_pairs):
        fig_ca_change = px.box(
            ca_pairs,
            x="subject",
            y="change",
            color="student_group",
            points=False,
            title="Distribution of CAASPP score change",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"subject": "CAASPP subject", "change": "Score change", "student_group": ""},
        )
        fig_ca_change = soften_axes(polish(fig_ca_change), x_title="", y_title="Score change", hide_repeated_y=False)

        ca_start_end = ca_pairs.melt(
            id_vars=["student_id", "student_group", "subject"],
            value_vars=["start_value", "end_value"],
            var_name="point",
            value_name="score",
        )
        ca_start_end["point"] = ca_start_end["point"].map({"start_value": "Start", "end_value": "End"})
        ca_line = ca_start_end.groupby(["student_group", "subject", "point"])["score"].mean().reset_index()
        fig_ca_level = px.line(
            ca_line,
            x="point",
            y="score",
            color="student_group",
            facet_col="subject",
            markers=True,
            title="Average CAASPP start-to-end movement",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"point": "", "score": "Avg score", "student_group": "", "subject": ""},
        )
        fig_ca_level = soften_axes(polish(fig_ca_level), x_title="", y_title="Avg CAASPP")
    else:
        fig_ca_change = empty_fig("Distribution of CAASPP score change")
        fig_ca_level = empty_fig("Average CAASPP start-to-end movement")

    att_pairs = pairs[pairs["metric"].eq("Attendance")]
    if len(att_pairs):
        fig_att_change = px.box(
            att_pairs,
            x="student_group",
            y="change",
            color="student_group",
            points=False,
            title="Distribution of attendance-rate change",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"student_group": "", "change": "Attendance point change"},
        )
        fig_att_change = soften_axes(polish(fig_att_change), x_title="", y_title="Point change", hide_repeated_y=False)
    else:
        fig_att_change = empty_fig("Distribution of attendance-rate change")

    relationship = pd.DataFrame()
    if len(att_pairs) and len(ca_pairs):
        ca_avg = ca_pairs.groupby(["student_id", "student_group", "student_name"], as_index=False)["change"].mean().rename(columns={"change": "caaspp_change"})
        att_avg = att_pairs[["student_id", "change"]].rename(columns={"change": "attendance_change"})
        relationship = ca_avg.merge(att_avg, on="student_id", how="inner")
    if len(relationship):
        fig_relation = px.scatter(
            relationship,
            x="attendance_change",
            y="caaspp_change",
            color="student_group",
            hover_data=["student_name"],
            title="Do attendance gains move with CAASPP gains?",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"attendance_change": "Attendance point change", "caaspp_change": "Avg CAASPP score change", "student_group": ""},
        )
        fig_relation.add_hline(y=0, line_dash="dot", line_color="#9aa6b2")
        fig_relation.add_vline(x=0, line_dash="dot", line_color="#9aa6b2")
        fig_relation = soften_axes(polish(fig_relation), x_title="Attendance point change", y_title="Avg CAASPP change", hide_repeated_y=False)
    else:
        fig_relation = empty_fig("Do attendance gains move with CAASPP gains?")

    if len(pairs):
        table_df = pairs.copy()
        table_df["change_display"] = table_df["change"].map(lambda x: f"{x:+.1f}")
        table_df["start_value_display"] = table_df["start_value"].map(lambda x: f"{x:,.1f}")
        table_df["end_value_display"] = table_df["end_value"].map(lambda x: f"{x:,.1f}")
        table = snapshot_table(table_df.sort_values(["outcome", "change"], ascending=[True, False]))
    else:
        table = snapshot_table(pd.DataFrame())

    cards = html.Div(
        className="kpi-grid",
        children=[
            kpi("Students with annual pairs", f"{paired_students:,}", "CAASPP or attendance start/end values"),
            kpi("CAASPP ELA improved", ca_ela[0], ca_ela[1]),
            kpi("CAASPP Math improved", ca_math[0], ca_math[1]),
            kpi("Attendance improved", att_kpi[0], att_kpi[1]),
        ],
    )
    note = html.Div(
        className="insight-band",
        children=[
            html.H3("How to read this page"),
            html.P("Each student contributes one CAASPP pair per subject and one attendance pair, comparing the earliest and latest selected points in the current filter view. The top chart shows the share of paired students who improved, regressed, or stayed flat. The middle charts show the size of those changes for CAASPP and attendance. The lower-right scatter checks whether students with stronger attendance gains also tended to gain on CAASPP."),
        ],
    )
    return page(
        "CAASPP and Attendance",
        "Earliest-to-latest CAASPP and attendance movement behind the overview story.",
        html.Div([cards, note, dcc.Graph(figure=fig_direction, className="chart wide"), html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca_change), dcc.Graph(figure=fig_att_change)]), html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca_level), dcc.Graph(figure=fig_relation)]), html.H2("CAASPP and attendance student change details"), html.Div(table, className="table-scroll")]),
    )


def student_page(dfs):
    students, star, caaspp, attendance = dfs["students"], dfs["star"], dfs["caaspp"], dfs["attendance"]
    star_plot = star.copy()
    if len(star_plot):
        star_plot["period_label"] = star_plot["school_year"] + " " + star_plot["season"]
        fig_star = px.line(
            star_plot,
            x="period_label",
            y="value",
            color="subject",
            line_group="student_id",
            markers=True,
            hover_data=["student_name", "student_id", "student_group"],
            title="Student STAR trajectory",
            template=TEMPLATE,
            color_discrete_map={"Reading": BLUE, "Math": GREEN},
            labels={"period_label": "Year and season", "value": "STAR value", "subject": ""},
        )
        fig_star = polish(fig_star)
        fig_star = soften_axes(fig_star, x_title="", y_title="STAR value", hide_repeated_y=False)
    else:
        fig_star = empty_fig("Student STAR trajectory")
    if len(caaspp):
        fig_ca = px.bar(
            caaspp,
            x="school_year_display",
            y="value",
            color="subject",
            barmode="group",
            hover_data=["student_name", "student_id", "student_group"],
            title="Student CAASPP records",
            template=TEMPLATE,
            labels={"school_year_display": "Year", "value": "CAASPP score", "subject": ""},
        )
        fig_ca = polish(fig_ca)
        fig_ca = soften_axes(fig_ca, x_title="", y_title="CAASPP score", hide_repeated_y=False)
    else:
        fig_ca = empty_fig("Student CAASPP records")
    if len(attendance):
        fig_att = px.bar(
            attendance,
            x="school_year",
            y="value",
            color="measure",
            hover_data=["student_name", "student_id", "student_group"],
            title="Student attendance records",
            template=TEMPLATE,
            labels={"school_year": "Year", "value": "Attendance rate", "measure": ""},
        )
        fig_att.update_yaxes(range=[0, 100])
        fig_att = polish(fig_att)
        fig_att = soften_axes(fig_att, x_title="", y_title="Attendance", hide_repeated_y=False)
    else:
        fig_att = empty_fig("Student attendance records")
    cols = ["student_id", "student_name", "student_group", "school_current", "grade_current", "ethnicity_group", "stepup_exposure_level", "intervention_intensity"]
    table = html.Table(
        className="data-table",
        children=[
            html.Thead(html.Tr([html.Th(c.replace("_", " ").title()) for c in cols])),
            html.Tbody([html.Tr([html.Td(str(row.get(c, ""))) for c in cols]) for _, row in students[cols].head(60).iterrows()]),
        ],
    )
    return page("Student Drilldown", "Search one or more students above to inspect the records we have for them.", html.Div([dcc.Graph(figure=fig_star, className="chart wide"), html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca), dcc.Graph(figure=fig_att)]), html.H2("Students in selection"), table]))


def dosage_page(dfs, dosage_programs=None, dosage_buckets=None):
    dosage = dfs.get("dosage", pd.DataFrame()).copy()
    if dosage.empty:
        return page(
            "Dosage vs Performance",
            "See how STAR, CAASPP, and attendance move as STEP UP dosage increases.",
            html.Div(
                className="insight-band",
                children=[
                    html.H3("No dosage data available"),
                    html.P("The dosage workbook did not produce any matched student records for the current filters."),
                ],
            ),
        )

    bucket_values = set(normalize_values(dosage_buckets))
    if bucket_values:
        dosage = dosage[dosage["dosage_bucket"].isin(bucket_values)]
    if dosage.empty:
        return page(
            "Dosage vs Performance",
            "See how STAR, CAASPP, and attendance move as STEP UP dosage increases.",
            html.Div(
                className="insight-band",
                children=[
                    html.H3("No dosage data for this selection"),
                    html.P("Try widening the dosage bucket filter or the global student filters."),
                ],
            ),
        )

    star = dfs["star"].copy()
    if len(star):
        star = star[star["value_type"].eq("score")].copy()
        star = star.merge(dosage[["student_id", "dosage_bucket"]], on="student_id", how="inner")
        star["season_order"] = star["season"].map(SEASON_ORDER).fillna(9)
        star["time_order"] = star["school_year"].astype(str).str.slice(0, 4).astype(int) * 10 + star["season_order"]
        star["time_label"] = star["school_year"].astype(str) + " " + star["season"].astype(str)
        star_trend = star.groupby(["time_label", "time_order", "subject", "dosage_bucket"])["value"].mean().reset_index().sort_values("time_order")
        fig_star = px.line(
            star_trend,
            x="time_label",
            y="value",
            color="dosage_bucket",
            facet_col="subject",
            markers=True,
            title="STAR trend over time by dosage bucket",
            template=TEMPLATE,
            color_discrete_map=DOSAGE_BUCKET_COLORS,
            category_orders={"dosage_bucket": DOSAGE_BUCKET_ORDER},
            labels={"time_label": "Year and season", "value": "Average STAR score", "dosage_bucket": "Dosage bucket", "subject": ""},
        )
        fig_star = polish(fig_star)
        fig_star = soften_axes(fig_star, x_title="", y_title="Average STAR score", hide_repeated_y=False)
    else:
        fig_star = empty_fig("STAR trend over time by dosage bucket")

    caaspp = dfs["caaspp"].copy()
    if len(caaspp):
        caaspp = caaspp.merge(dosage[["student_id", "dosage_bucket"]], on="student_id", how="inner")
        ca_trend = caaspp.groupby(["school_year_display", "subject", "dosage_bucket"])["value"].mean().reset_index().sort_values("school_year_display")
        fig_ca = px.line(
            ca_trend,
            x="school_year_display",
            y="value",
            color="dosage_bucket",
            facet_col="subject",
            markers=True,
            title="CAASPP trend over time by dosage bucket",
            template=TEMPLATE,
            color_discrete_map=DOSAGE_BUCKET_COLORS,
            category_orders={"dosage_bucket": DOSAGE_BUCKET_ORDER},
            labels={"school_year_display": "Year", "value": "Average CAASPP score", "dosage_bucket": "Dosage bucket", "subject": ""},
        )
        fig_ca = polish(fig_ca)
        fig_ca = soften_axes(fig_ca, x_title="", y_title="Average CAASPP score", hide_repeated_y=False)
    else:
        fig_ca = empty_fig("CAASPP trend over time by dosage bucket")

    attendance = dfs["attendance"].copy()
    if len(attendance):
        attendance["measure_rank"] = attendance["measure"].map({"SIS Reported Rate": 0, "Attendance rate": 1, "Full Day Rate": 2}).fillna(9)
        attendance = attendance.sort_values("measure_rank").drop_duplicates(["student_id", "school_year"], keep="first")
        attendance = attendance.merge(dosage[["student_id", "dosage_bucket"]], on="student_id", how="inner")
        att_trend = attendance.groupby(["school_year", "dosage_bucket"])["value"].mean().reset_index().sort_values("school_year")
        fig_att = px.line(
            att_trend,
            x="school_year",
            y="value",
            color="dosage_bucket",
            markers=True,
            title="Attendance trend over time by dosage bucket",
            template=TEMPLATE,
            color_discrete_map=DOSAGE_BUCKET_COLORS,
            category_orders={"dosage_bucket": DOSAGE_BUCKET_ORDER},
            labels={"school_year": "Year", "value": "Average attendance", "dosage_bucket": "Dosage bucket"},
        )
        fig_att.update_yaxes(range=[0, 100])
        fig_att = polish(fig_att)
        fig_att = soften_axes(fig_att, x_title="", y_title="Average attendance", hide_repeated_y=False)
    else:
        fig_att = empty_fig("Attendance trend over time by dosage bucket")

    bucket_profile = dosage.groupby("dosage_bucket")["student_id"].nunique().reindex(DOSAGE_BUCKET_ORDER, fill_value=0).reset_index(name="students")
    fig_bucket = px.bar(
        bucket_profile,
        x="dosage_bucket",
        y="students",
        color="dosage_bucket",
        title="Students by dosage bucket",
        template=TEMPLATE,
        color_discrete_map=DOSAGE_BUCKET_COLORS,
        category_orders={"dosage_bucket": DOSAGE_BUCKET_ORDER},
        labels={"dosage_bucket": "Dosage bucket", "students": "Students"},
    )
    fig_bucket = polish(fig_bucket)
    fig_bucket = soften_axes(fig_bucket, x_title="", y_title="Students", hide_repeated_y=False)

    summary = dosage.sort_values(["dosage_sessions_attended", "student_name"], ascending=[False, True]).copy()
    summary["dosage_rate_display"] = summary["dosage_rate"].map(lambda x: f"{x * 100:.1f}%")
    summary["dosage_sessions_attended"] = summary["dosage_sessions_attended"].astype(int)
    summary["dosage_sessions_possible"] = summary["dosage_sessions_possible"].astype(int)
    table_cols = [
        "student_name",
        "school_current",
        "grade_current",
        "student_group",
        "dosage_sessions_attended",
        "dosage_sessions_possible",
        "dosage_rate_display",
        "dosage_bucket",
        "program_count",
    ]
    table_labels = ["Student", "School", "Grade", "Group", "Attended", "Possible", "Dosage rate", "Bucket", "Programs"]
    table = html.Table(
        className="data-table compact-table",
        children=[
            html.Thead(html.Tr([header_cell(label) for label in table_labels])),
            html.Tbody([html.Tr([html.Td(str(row.get(col, ""))) for col in table_cols]) for _, row in summary[table_cols].head(60).iterrows()]),
        ],
    )

    student_count = summary["student_id"].nunique()
    avg_sessions = summary["dosage_sessions_attended"].mean()
    avg_rate = summary["dosage_rate"].mean()
    median_sessions = summary["dosage_sessions_attended"].median()
    cards = html.Div(
        className="kpi-grid",
        children=[
            kpi("Students with dosage", f"{student_count:,}", "Matched STEP UP students with attendance records"),
            kpi("Avg attended sessions", f"{avg_sessions:.1f}", "Across matched dosage records"),
            kpi("Average dosage rate", f"{avg_rate * 100:.1f}%", "Attended sessions divided by possible sessions"),
            kpi("Median attended sessions", f"{median_sessions:.1f}", "Middle dosage value in the filtered view"),
        ],
    )
    note = html.Div(
        className="insight-band",
        children=[
            html.H3("How to read this page"),
            html.P(
                "Dosage here is the count of attended sessions across the STEP UP attendance workbook, normalized into comparable percentage buckets because the underlying programs have different session counts. "
                "Use the dosage bucket filter to narrow the view, then compare how STAR, CAASPP, and attendance move over time for students in that dose range."
            ),
        ],
    )
    return page(
        "Dosage vs Performance",
        "See how STAR, CAASPP, and attendance move as STEP UP dosage increases.",
        html.Div(
            [
                cards,
                note,
                dcc.Graph(figure=fig_star, className="chart wide"),
                html.Div(className="two-col", children=[dcc.Graph(figure=fig_ca), dcc.Graph(figure=fig_att)]),
                dcc.Graph(figure=fig_bucket, className="chart wide"),
                html.H2("Dosage student detail"),
                html.Div(table, className="table-scroll"),
            ]
        ),
    )


def coverage_page(dfs):
    students = dfs["students"]
    total_students = students["student_id"].nunique()
    step_students = students.loc[students["student_group"].eq("STEP UP"), "student_id"].nunique()
    non_students = students.loc[students["student_group"].eq("Non-STEP UP"), "student_id"].nunique()
    complete_students = int(students["strict_comparison_ready"].sum()) if "strict_comparison_ready" in students else 0

    availability_rows = []
    star = dfs["star"]
    if len(star):
        star_counts = star.groupby(["student_group", "school_year", "season", "subject"])["student_id"].nunique().reset_index(name="students")
        for _, row in star_counts.iterrows():
            availability_rows.append(
                {
                    "Metric": "STAR",
                    "Year": row["school_year"],
                    "Session": row["season"],
                    "Subject": row["subject"],
                    "Group": row["student_group"],
                    "Students": int(row["students"]),
                    "Confidence": "High",
                }
            )
    caaspp = dfs["caaspp"]
    if len(caaspp):
        ca_counts = caaspp.groupby(["student_group", "school_year_display", "subject"])["student_id"].nunique().reset_index(name="students")
        for _, row in ca_counts.iterrows():
            confidence = "Medium" if str(row["school_year_display"]) == "2023-24" else "High"
            availability_rows.append(
                {
                    "Metric": "CAASPP",
                    "Year": row["school_year_display"],
                    "Session": "Yearly",
                    "Subject": row["subject"],
                    "Group": row["student_group"],
                    "Students": int(row["students"]),
                    "Confidence": confidence,
                }
            )
    attendance = dfs["attendance"]
    if len(attendance):
        att_counts = attendance.groupby(["student_group", "school_year"])["student_id"].nunique().reset_index(name="students")
        for _, row in att_counts.iterrows():
            confidence = "High" if str(row["school_year"]) == "2024-25" else "Medium"
            availability_rows.append(
                {
                    "Metric": "Attendance",
                    "Year": row["school_year"],
                    "Session": "Annual",
                    "Subject": "Attendance",
                    "Group": row["student_group"],
                    "Students": int(row["students"]),
                    "Confidence": confidence,
                }
            )
    matrix = pd.DataFrame(availability_rows)
    if len(matrix):
        matrix["Coverage"] = matrix["Metric"] + " | " + matrix["Subject"] + " | " + matrix["Year"] + " | " + matrix["Session"]
        fig_matrix = px.bar(
            matrix,
            x="Students",
            y="Coverage",
            color="Group",
            barmode="group",
            orientation="h",
            title="Coverage matrix: metric, year, session, and student group",
            template=TEMPLATE,
            color_discrete_map={"STEP UP": GREEN, "Non-STEP UP": GOLD},
            labels={"Students": "Students with data", "Coverage": "", "Group": ""},
            hover_data=["Confidence"],
        )
        fig_matrix.update_layout(height=max(520, min(1050, 26 * matrix["Coverage"].nunique() + 140)))
        fig_matrix = polish(fig_matrix)
        fig_matrix = soften_axes(fig_matrix, x_title="Students", y_title="", hide_repeated_y=False)
    else:
        fig_matrix = empty_fig("Coverage matrix")

    expected = []
    for metric, years, sessions, subjects in [
        ("STAR", ["2022-23", "2023-24", "2024-25", "2025-26"], ["Fall", "Winter", "Spring"], ["Reading", "Math"]),
        ("CAASPP", ["2022-23", "2023-24", "2024-25", "2025-26"], ["Yearly"], ["ELA", "Math"]),
        ("Attendance", ["2022-23", "2023-24", "2024-25", "2025-26"], ["Annual"], ["Attendance"]),
    ]:
        for year in years:
            for session in sessions:
                for subject in subjects:
                    expected.append({"Metric": metric, "Year": year, "Session": session, "Subject": subject})
    expected_df = pd.DataFrame(expected)
    present = matrix.groupby(["Metric", "Year", "Session", "Subject"])["Students"].sum().reset_index() if len(matrix) else pd.DataFrame(columns=["Metric", "Year", "Session", "Subject", "Students"])
    present["Status"] = "Present"
    gap = expected_df.merge(present, on=["Metric", "Year", "Session", "Subject"], how="left")
    gap["Status"] = gap["Status"].fillna("Missing")
    gap["Students"] = gap["Students"].fillna(0).astype(int)
    gap["Field"] = gap["Metric"] + " | " + gap["Subject"] + " | " + gap["Session"]
    fig_gap = px.bar(
        gap,
        x="Year",
        y="Field",
        color="Status",
        orientation="h",
        title="Present vs missing data slots",
        template=TEMPLATE,
        color_discrete_map={"Present": GREEN, "Missing": "#c9d1da"},
        labels={"Year": "Year", "Field": "", "Status": ""},
        hover_data={"Students": True},
    )
    fig_gap = polish(fig_gap)
    fig_gap = soften_axes(fig_gap, x_title="", y_title="", hide_repeated_y=False)

    table_df = matrix.sort_values(["Metric", "Year", "Session", "Subject", "Group"]) if len(matrix) else matrix
    table = html.Table(
        className="data-table compact-table",
        children=[
            html.Thead(html.Tr([header_cell(c, COVERAGE_COLUMN_HELP.get(c)) for c in ["Metric", "Year", "Session", "Subject", "Group", "Students", "Confidence"]])),
            html.Tbody(
                [
                    html.Tr([html.Td(str(row.get(c, ""))) for c in ["Metric", "Year", "Session", "Subject", "Group", "Students", "Confidence"]])
                    for _, row in table_df.iterrows()
                ]
            ),
        ],
    )
    fig_star_avail, fig_avail = availability_figures(dfs)
    cards = html.Div(
        className="kpi-grid",
        children=[
            kpi("Students in selection", f"{total_students:,}", "Current filters"),
            kpi("STEP UP students", f"{step_students:,}", "Data-file STEP UP evidence"),
            kpi("Non-STEP UP students", f"{non_students:,}", "No STEP UP evidence"),
            kpi("Complete comparison students", f"{complete_students:,}", "All core fields available"),
        ],
    )
    note = html.Div(
        className="insight-band",
        children=[
            html.H3("How to read coverage"),
            html.P("This page is only about data availability. STAR is shown by year, season, and subject. CAASPP is yearly by subject. Attendance is annual. Medium-confidence fields are usable but have inferred-year context documented in the data notes. Strict mode stays on the 2022-25 comparison years for now, so 2025-26 appears only in All available mode until CAASPP arrives."),
        ],
    )
    return page(
        "Data Coverage",
        "Which STAR, CAASPP, and attendance fields are present or missing by year, session, subject, and student group.",
        html.Div(
            [
                cards,
                note,
                dcc.Graph(figure=fig_gap, className="chart wide"),
                dcc.Graph(figure=fig_matrix, className="chart wide"),
                dcc.Graph(figure=fig_star_avail, className="chart wide"),
                dcc.Graph(figure=fig_avail, className="chart wide"),
                html.H2("Coverage detail"),
                html.Div(table, className="table-scroll"),
            ]
        ),
    )


app = Dash(__name__, suppress_callback_exceptions=True, title="STEP UP Student Outcomes")
server = app.server
app.layout = base_layout


@app.callback(
    Output("student-filter", "options"),
    Output("subject-filter", "options"),
    Output("year-filter", "options"),
    Output("period-filter", "options"),
    Output("school-filter", "options"),
    Output("grade-filter", "options"),
    Output("ethnicity-filter", "options"),
    Input("analysis-mode", "value"),
    Input("group-filter", "value"),
    Input("subject-filter", "value"),
    Input("year-filter", "value"),
    Input("period-filter", "value"),
    Input("school-filter", "value"),
    Input("grade-filter", "value"),
    Input("ethnicity-filter", "value"),
    Input("student-filter", "value"),
)
def cascade_filter_options(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    students = DS["students"].copy()
    if mode == "strict":
        students = students[students["strict_comparison_ready"].astype(bool)]
    if group != "all":
        students = students[students["student_group"].eq(group)]
    if student_ids:
        students = students[students["student_id"].isin([str(x) for x in student_ids])]
    if schools:
        students = students[students["school_group"].isin(schools)]
    if grades:
        students = students[students["grade_current"].isin(grades)]
    if ethnicities:
        students = students[students["ethnicity_group"].isin(ethnicities)]

    ids = set(students["student_id"])
    star = DS["star"][DS["star"]["student_id"].isin(ids)]
    growth = DS["growth"][DS["growth"]["student_id"].isin(ids)]
    caaspp = DS["caaspp"][DS["caaspp"]["student_id"].isin(ids)]
    subject_context = pd.concat(
        [star["subject"], caaspp["subject"].replace({"ELA": "Reading"})],
        ignore_index=True,
    )
    if subjects:
        star = star[star["subject"].isin(subjects)]
        growth = growth[growth["subject"].isin(subjects)]
        ca_subjects = set(subjects)
        if "Reading" in ca_subjects:
            ca_subjects.add("ELA")
        caaspp = caaspp[caaspp["subject"].isin(ca_subjects)]
    if years:
        star = star[star["school_year"].isin(years)]
        growth = growth[growth["school_year"].isin(years)]
        caaspp = caaspp[caaspp["school_year_display"].isin(years)]
    if periods:
        growth = growth[growth["period"].isin(periods)]

    dosage_year_values = []
    if len(DS["dosage"]):
        dosage_subset = DS["dosage"][DS["dosage"]["student_id"].isin(ids)]
        if "program_years" in dosage_subset.columns:
            for text in dosage_subset["program_years"].dropna().astype(str):
                dosage_year_values.extend([part.strip() for part in text.split(";") if part.strip()])

    student_options = opts_from_frame(students.sort_values("student_label"), "student_label", "student_id")
    subject_options = opts(subject_context[subject_context.isin(["Reading", "Math"])])
    year_values = pd.concat(
        [
            star["school_year"],
            growth["school_year"],
            caaspp["school_year_display"],
            DS["attendance"][DS["attendance"]["student_id"].isin(ids)]["school_year"],
            pd.Series(dosage_year_values, dtype=object),
        ],
        ignore_index=True,
    )
    year_options = opts(year_values)
    period_options = opts(growth["period"])
    school_options = opts(students["school_group"])
    grade_options = opts(students["grade_current"].dropna().astype(int))
    ethnicity_options = opts(students["ethnicity_group"])
    return student_options, subject_options, year_options, period_options, school_options, grade_options, ethnicity_options


@app.callback(
    Output("overview-snapshot-table", "children"),
    Input("overview-outcome-filter", "value"),
    Input("analysis-mode", "value"),
    Input("group-filter", "value"),
    Input("subject-filter", "value"),
    Input("year-filter", "value"),
    Input("period-filter", "value"),
    Input("school-filter", "value"),
    Input("grade-filter", "value"),
    Input("ethnicity-filter", "value"),
    Input("student-filter", "value"),
)
def update_overview_snapshot(outcome, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    df = filtered_snapshot(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, outcome)
    return snapshot_table(df)


@app.callback(
    Output("download-snapshot", "data"),
    Input("download-snapshot-button", "n_clicks"),
    State("overview-outcome-filter", "value"),
    State("analysis-mode", "value"),
    State("group-filter", "value"),
    State("subject-filter", "value"),
    State("year-filter", "value"),
    State("period-filter", "value"),
    State("school-filter", "value"),
    State("grade-filter", "value"),
    State("ethnicity-filter", "value"),
    State("student-filter", "value"),
    prevent_initial_call=True,
)
def download_overview_snapshot(n_clicks, outcome, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids):
    if not n_clicks:
        raise PreventUpdate
    df = filtered_snapshot(mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, outcome)
    if df.empty:
        raise PreventUpdate
    filename = f"student_outcome_snapshot_{str(outcome or 'outcome').lower().replace(' ', '_')}.xlsx"
    return dcc.send_bytes(snapshot_excel_bytes(df), filename)


@app.callback(
    Output("summary-content", "children"),
    Input("summary-metric-filter", "value"),
    Input("analysis-mode", "value"),
    Input("group-filter", "value"),
    Input("year-filter", "value"),
    Input("school-filter", "value"),
    Input("grade-filter", "value"),
    Input("ethnicity-filter", "value"),
    Input("student-filter", "value"),
)
def update_summary_content(metric, mode, group, years, schools, grades, ethnicities, student_ids):
    dfs = summary_filtered_dfs(mode, group, years, schools, grades, ethnicities, student_ids)
    rows = summary_metric_rows(dfs, metric)
    return summary_cards(rows, metric, years)


@app.callback(
    Output("download-summary", "data"),
    Input("download-summary-button", "n_clicks"),
    State("summary-metric-filter", "value"),
    State("analysis-mode", "value"),
    State("group-filter", "value"),
    State("year-filter", "value"),
    State("school-filter", "value"),
    State("grade-filter", "value"),
    State("ethnicity-filter", "value"),
    State("student-filter", "value"),
    prevent_initial_call=True,
)
def download_summary(n_clicks, metric, mode, group, years, schools, grades, ethnicities, student_ids):
    if not n_clicks:
        raise PreventUpdate
    dfs = summary_filtered_dfs(mode, group, years, schools, grades, ethnicities, student_ids)
    rows = summary_metric_rows(dfs, metric)
    if rows.empty:
        raise PreventUpdate
    filename = f"step_up_summary_{str(metric or 'metric').lower().replace(' ', '_')}.xlsx"
    return dcc.send_bytes(summary_excel_bytes(rows), filename)


@app.callback(
    Output("page", "children"),
    Input("url", "pathname"),
    Input("analysis-mode", "value"),
    Input("group-filter", "value"),
    Input("subject-filter", "value"),
    Input("year-filter", "value"),
    Input("period-filter", "value"),
    Input("school-filter", "value"),
    Input("grade-filter", "value"),
    Input("ethnicity-filter", "value"),
    Input("student-filter", "value"),
    Input("dosage-bucket-filter", "value"),
)
def render(pathname, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, dosage_buckets):
    try:
        return cached_render_from_key(page_cache_key(pathname, mode, group, subjects, years, periods, schools, grades, ethnicities, student_ids, None, dosage_buckets))
    except Exception as exc:
        import traceback

        print("Dashboard render error:", repr(exc))
        traceback.print_exc()
        return error_panel(
            "The page could not be rendered on the server. Check the Render logs for the traceback, then I can fix the underlying data or callback error."
        )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8050"))
    app.run(debug=False, host="0.0.0.0", port=port)
